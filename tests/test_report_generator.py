"""Unit tests for valuation.report_generator and the new FinancialDataFetcher
methods get_current_price / get_historical_fcf_growth.

All tests are fully offline – no real HTTP calls are made.
"""

from __future__ import annotations

import os
import tempfile
import unittest
from unittest.mock import MagicMock, call, patch

# ---------------------------------------------------------------------------
# Helpers shared with the existing test_data_fetcher convention
# ---------------------------------------------------------------------------

_DUMMY_KEY = "test_api_key"
_TICKER = "AAPL"


def _mock_response(json_data, status_code: int = 200) -> MagicMock:
    mock_resp = MagicMock()
    mock_resp.status_code = status_code
    mock_resp.json.return_value = json_data
    if status_code >= 400:
        import requests

        mock_resp.raise_for_status.side_effect = requests.exceptions.HTTPError(
            response=mock_resp
        )
    else:
        mock_resp.raise_for_status.return_value = None
    return mock_resp


# ---------------------------------------------------------------------------
# FinancialDataFetcher.get_current_price
# ---------------------------------------------------------------------------

from valuation.data_fetcher import DataFetchError, FinancialDataFetcher


class TestGetCurrentPrice(unittest.TestCase):
    _SAMPLE = [{"symbol": "AAPL", "price": 189.84, "marketCap": 2_900_000_000_000}]

    @patch("valuation.data_fetcher.requests.get")
    def test_returns_price_and_symbol(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response(self._SAMPLE)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_current_price()
        self.assertAlmostEqual(result["price"], 189.84)
        self.assertEqual(result["symbol"], "AAPL")

    @patch("valuation.data_fetcher.requests.get")
    def test_price_converted_to_float(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response([{"symbol": "AAPL", "price": 200}])
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_current_price()
        self.assertIsInstance(result["price"], float)

    @patch("valuation.data_fetcher.requests.get")
    def test_empty_list_raises(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response([])
        with self.assertRaises(DataFetchError):
            FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY).get_current_price()

    @patch("valuation.data_fetcher.requests.get")
    def test_missing_price_field_raises(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response([{"symbol": "AAPL"}])
        with self.assertRaises(DataFetchError):
            FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY).get_current_price()

    @patch("valuation.data_fetcher.requests.get")
    def test_http_error_raises(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response({}, status_code=401)
        with self.assertRaises(DataFetchError):
            FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY).get_current_price()


# ---------------------------------------------------------------------------
# FinancialDataFetcher.get_historical_fcf_growth
# ---------------------------------------------------------------------------

# Five years of FCF data, newest-first (as FMP returns)
_FIVE_YEAR_CF = [
    {"date": "2023-09-30", "freeCashFlow": 99_584_000_000},
    {"date": "2022-09-30", "freeCashFlow": 111_443_000_000},
    {"date": "2021-09-30", "freeCashFlow": 92_953_000_000},
    {"date": "2020-09-30", "freeCashFlow": 73_365_000_000},
    {"date": "2019-09-30", "freeCashFlow": 58_896_000_000},
]


class TestGetHistoricalFcfGrowth(unittest.TestCase):
    @patch("valuation.data_fetcher.requests.get")
    def test_returns_float(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response(_FIVE_YEAR_CF)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        self.assertIsInstance(result, float)

    @patch("valuation.data_fetcher.requests.get")
    def test_cagr_uses_oldest_and_latest(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response(_FIVE_YEAR_CF)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        # Records reversed → oldest = 58_896M, latest = 99_584M, n=4
        expected_cagr = (99_584_000_000 / 58_896_000_000) ** (1 / 4) - 1
        self.assertAlmostEqual(result, expected_cagr, places=8)

    @patch("valuation.data_fetcher.requests.get")
    def test_single_record_returns_default(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response([{"date": "2023-09-30", "freeCashFlow": 100}])
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        self.assertAlmostEqual(result, 0.05)

    @patch("valuation.data_fetcher.requests.get")
    def test_empty_list_returns_default(self, mock_get: MagicMock) -> None:
        mock_get.return_value = _mock_response([])
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        self.assertAlmostEqual(result, 0.05)

    @patch("valuation.data_fetcher.requests.get")
    def test_all_negative_fcf_returns_default(self, mock_get: MagicMock) -> None:
        records = [
            {"date": "2023-09-30", "freeCashFlow": -500},
            {"date": "2022-09-30", "freeCashFlow": -400},
            {"date": "2021-09-30", "freeCashFlow": -300},
        ]
        mock_get.return_value = _mock_response(records)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        self.assertAlmostEqual(result, 0.05)

    @patch("valuation.data_fetcher.requests.get")
    def test_partial_positive_uses_average_yoy(self, mock_get: MagicMock) -> None:
        # oldest=-100 (invalid), then two positive consecutive values
        records = [
            {"date": "2023-09-30", "freeCashFlow": 120},  # newest
            {"date": "2022-09-30", "freeCashFlow": 100},
            {"date": "2021-09-30", "freeCashFlow": -50},  # negative → no CAGR
        ]
        mock_get.return_value = _mock_response(records)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        # Only valid YoY: 100→120 = 20%
        self.assertAlmostEqual(result, 0.20, places=8)

    @patch("valuation.data_fetcher.requests.get")
    def test_two_positive_records_gives_cagr(self, mock_get: MagicMock) -> None:
        records = [
            {"date": "2023-09-30", "freeCashFlow": 121},
            {"date": "2022-09-30", "freeCashFlow": 100},
        ]
        mock_get.return_value = _mock_response(records)
        fetcher = FinancialDataFetcher(_TICKER, api_key=_DUMMY_KEY)
        result = fetcher.get_historical_fcf_growth()
        self.assertAlmostEqual(result, 0.21, places=8)


# ---------------------------------------------------------------------------
# run_scenario_analysis – integration (fully mocked, no real HTTP or disk I/O)
# ---------------------------------------------------------------------------

from valuation.report_generator import _SCENARIO_ORDER, run_scenario_analysis

# Canonical mock data
_CF_DATA = {
    "freeCashFlow": 100_000_000_000,
    "operatingCashFlow": 120_000_000_000,
    "date": "2023-09-30",
}
_EM_DATA = {
    "totalDebt": 111_000_000_000,
    "cashAndCashEquivalents": 30_000_000_000,
    "sharesOutstanding": 15_600_000_000,
    "date": "2023-09-30",
}
_WACC_DATA = {
    "wacc": 0.09,
    "costOfEquity": 0.10,
    "costOfDebt": 0.03,
    "taxRate": 0.15,
    "equityWeight": 0.92,
    "debtWeight": 0.08,
}
_PRICE_DATA = {"price": 189.84, "symbol": "AAPL"}
_BASE_GROWTH = 0.08  # historical FCF CAGR


def _make_mock_fetcher(
    cf_data=None,
    em_data=None,
    wacc_data=None,
    price_data=None,
    base_growth=None,
) -> MagicMock:
    fetcher = MagicMock(spec=FinancialDataFetcher)
    fetcher.get_cash_flow_statement.return_value = cf_data or _CF_DATA
    fetcher.get_enterprise_metrics.return_value = em_data or _EM_DATA
    fetcher.get_wacc.return_value = wacc_data or _WACC_DATA
    fetcher.get_current_price.return_value = price_data or _PRICE_DATA
    fetcher.get_historical_fcf_growth.return_value = (
        base_growth if base_growth is not None else _BASE_GROWTH
    )
    return fetcher


class TestRunScenarioAnalysis(unittest.TestCase):
    """End-to-end tests with mocked FinancialDataFetcher and temp output dir."""

    def _run(self, ticker: str = _TICKER, **fetcher_overrides) -> str:
        mock_fetcher = _make_mock_fetcher(**fetcher_overrides)
        with tempfile.TemporaryDirectory() as tmpdir:
            with (
                patch(
                    "valuation.report_generator.FinancialDataFetcher",
                    return_value=mock_fetcher,
                ),
                patch(
                    "valuation.report_generator.os.makedirs"
                ),
                patch(
                    "valuation.report_generator.os.path.join",
                    side_effect=lambda *args: os.path.join(*args),
                ),
                patch(
                    "valuation.report_generator.os.path.dirname",
                    side_effect=lambda p: os.path.dirname(p),
                ),
                patch(
                    "valuation.report_generator.os.path.abspath",
                    side_effect=lambda p: os.path.abspath(p),
                ),
            ):
                # Override the output path to write into the temp dir
                excel_path = os.path.join(tmpdir, f"{ticker}_DCF_Valuation.xlsx")
                with patch(
                    "valuation.report_generator._build_excel"
                ) as mock_build:
                    result_path = run_scenario_analysis(ticker, api_key=_DUMMY_KEY)
                    return result_path, mock_build

    def test_returns_string_path(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with tempfile.TemporaryDirectory() as tmpdir:
            with (
                patch(
                    "valuation.report_generator.FinancialDataFetcher",
                    return_value=mock_fetcher,
                ),
                patch("valuation.report_generator.os.makedirs"),
                patch("valuation.report_generator._build_excel"),
            ):
                path = run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)
        self.assertIsInstance(path, str)
        self.assertIn(_TICKER, path)
        self.assertTrue(path.endswith(".xlsx"))

    def test_ticker_normalised_to_upper(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel"),
        ):
            path = run_scenario_analysis("aapl", api_key=_DUMMY_KEY)
        self.assertIn("AAPL", path)

    def test_all_fetcher_methods_called(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel"),
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        mock_fetcher.get_cash_flow_statement.assert_called_once()
        mock_fetcher.get_enterprise_metrics.assert_called_once()
        mock_fetcher.get_wacc.assert_called_once()
        mock_fetcher.get_current_price.assert_called_once()
        mock_fetcher.get_historical_fcf_growth.assert_called_once()

    def test_output_dir_created(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch(
                "valuation.report_generator.os.makedirs"
            ) as mock_makedirs,
            patch("valuation.report_generator._build_excel"),
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        # os.makedirs must have been called with exist_ok=True
        self.assertTrue(mock_makedirs.called)
        _, kwargs = mock_makedirs.call_args
        self.assertTrue(kwargs.get("exist_ok"))

    def test_build_excel_called_with_correct_args(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        self.assertTrue(mock_build.called)
        args, kwargs = mock_build.call_args
        path, ticker, current_price, scenarios, results = args
        self.assertEqual(ticker, _TICKER)
        self.assertAlmostEqual(current_price, _PRICE_DATA["price"])

    def test_three_scenarios_generated(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        args, _ = mock_build.call_args
        _, _, _, scenarios, results = args
        self.assertEqual(set(scenarios.keys()), {"Bear Case", "Base Case", "Bull Case"})
        self.assertEqual(set(results.keys()), {"Bear Case", "Base Case", "Bull Case"})

    def test_bull_growth_higher_than_base(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        _, _, _, scenarios, _ = mock_build.call_args[0]
        self.assertGreater(
            scenarios["Bull Case"]["fcf_growth_rate"],
            scenarios["Base Case"]["fcf_growth_rate"],
        )
        self.assertGreater(
            scenarios["Base Case"]["fcf_growth_rate"],
            scenarios["Bear Case"]["fcf_growth_rate"],
        )

    def test_bull_wacc_lower_than_base(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        _, _, _, scenarios, _ = mock_build.call_args[0]
        self.assertLess(
            scenarios["Bull Case"]["wacc"], scenarios["Base Case"]["wacc"]
        )
        self.assertLess(
            scenarios["Base Case"]["wacc"], scenarios["Bear Case"]["wacc"]
        )

    def test_scenario_deltas_are_correct(self) -> None:
        mock_fetcher = _make_mock_fetcher(base_growth=0.08)
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        _, _, _, scenarios, _ = mock_build.call_args[0]
        self.assertAlmostEqual(scenarios["Base Case"]["fcf_growth_rate"], 0.08)
        self.assertAlmostEqual(scenarios["Bull Case"]["fcf_growth_rate"], 0.10)
        self.assertAlmostEqual(scenarios["Bear Case"]["fcf_growth_rate"], 0.06)
        self.assertAlmostEqual(scenarios["Base Case"]["wacc"], 0.09)
        self.assertAlmostEqual(scenarios["Bull Case"]["wacc"], 0.08)
        self.assertAlmostEqual(scenarios["Bear Case"]["wacc"], 0.10)

    def test_bull_intrinsic_value_highest(self) -> None:
        mock_fetcher = _make_mock_fetcher()
        with (
            patch(
                "valuation.report_generator.FinancialDataFetcher",
                return_value=mock_fetcher,
            ),
            patch("valuation.report_generator.os.makedirs"),
            patch("valuation.report_generator._build_excel") as mock_build,
        ):
            run_scenario_analysis(_TICKER, api_key=_DUMMY_KEY)

        _, _, _, _, results = mock_build.call_args[0]
        bear_iv = results["Bear Case"]["intrinsic_value_per_share"]
        base_iv = results["Base Case"]["intrinsic_value_per_share"]
        bull_iv = results["Bull Case"]["intrinsic_value_per_share"]
        self.assertGreater(bull_iv, base_iv)
        self.assertGreater(base_iv, bear_iv)


# ---------------------------------------------------------------------------
# _build_excel – writes an actual file and checks sheet structure
# ---------------------------------------------------------------------------

from valuation.report_generator import _build_excel


class TestBuildExcelOutput(unittest.TestCase):
    """Smoke-tests that verify _build_excel writes a valid, well-structured
    Excel workbook without error."""

    def setUp(self) -> None:
        from valuation.dcf_model import DCFModel

        cf_data = {"freeCashFlow": 100_000_000_000}
        em_data = {
            "totalDebt": 50_000_000_000,
            "cashAndCashEquivalents": 20_000_000_000,
            "sharesOutstanding": 15_000_000_000,
        }
        model = DCFModel(cf_data, em_data)
        base_growth = 0.08
        wacc = 0.09

        self.scenarios = {
            "Bear Case": {"fcf_growth_rate": base_growth - 0.02, "wacc": wacc + 0.01},
            "Base Case": {"fcf_growth_rate": base_growth, "wacc": wacc},
            "Bull Case": {"fcf_growth_rate": base_growth + 0.02, "wacc": wacc - 0.01},
        }
        self.results = {}
        for name, params in self.scenarios.items():
            r = model.calculate_intrinsic_value(
                fcf_growth_rate=params["fcf_growth_rate"],
                wacc=params["wacc"],
            )
            r["wacc_used"] = params["wacc"]
            r["fcf_growth_rate"] = params["fcf_growth_rate"]
            self.results[name] = r

        self.current_price = 189.84
        self.tmpdir = tempfile.mkdtemp()
        self.excel_path = os.path.join(self.tmpdir, "AAPL_DCF_Valuation.xlsx")

    def tearDown(self) -> None:
        import shutil

        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _build(self) -> None:
        _build_excel(
            self.excel_path,
            "AAPL",
            self.current_price,
            self.scenarios,
            self.results,
        )

    def test_file_created(self) -> None:
        self._build()
        self.assertTrue(os.path.isfile(self.excel_path))

    def test_workbook_has_four_sheets(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        self.assertEqual(set(wb.sheetnames), {"Summary", "Bear Case", "Base Case", "Bull Case"})
        wb.close()

    def test_summary_sheet_has_title(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb["Summary"]
        title_cell = ws["A1"].value
        self.assertIsNotNone(title_cell)
        self.assertIn("AAPL", title_cell)
        wb.close()

    def test_scenario_sheets_have_fcf_data(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        for name in _SCENARIO_ORDER:
            ws = wb[name]
            # Row 5 should be the pandas-written FCF header
            header_val = ws.cell(row=5, column=1).value
            self.assertIsNotNone(header_val, f"Header missing in sheet '{name}'")
            # Row 6 should be year 1 data
            year_val = ws.cell(row=6, column=1).value
            self.assertEqual(year_val, 1, f"Year 1 missing in sheet '{name}'")
        wb.close()

    def test_summary_sheet_has_scenario_labels(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb["Summary"]
        # Row 5 should have scenario column headers
        values = [ws.cell(row=5, column=c).value for c in range(2, 5)]
        self.assertIn("Bear Case", values)
        self.assertIn("Base Case", values)
        self.assertIn("Bull Case", values)
        wb.close()

    def test_intrinsic_value_in_summary(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb["Summary"]
        # Scan column A for "Intrinsic Value / Share" label
        found = any(
            ws.cell(row=r, column=1).value
            and "Intrinsic" in str(ws.cell(row=r, column=1).value)
            for r in range(1, 25)
        )
        self.assertTrue(found, "Intrinsic Value / Share label not found in Summary sheet")
        wb.close()

    def test_valuation_bridge_in_scenario_sheet(self) -> None:
        self._build()
        from openpyxl import load_workbook

        wb = load_workbook(self.excel_path, read_only=True)
        ws = wb["Base Case"]
        # Row 12 should be the "VALUATION BRIDGE" header
        bridge_cell = ws.cell(row=12, column=1).value
        self.assertIsNotNone(bridge_cell)
        self.assertIn("VALUATION BRIDGE", str(bridge_cell).upper())
        wb.close()


if __name__ == "__main__":
    unittest.main()
