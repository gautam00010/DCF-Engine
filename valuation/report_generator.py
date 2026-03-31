"""Output generator for the DCF valuation pipeline.

Provides :func:`run_scenario_analysis`, which fetches live financial data,
runs Bear / Base / Bull DCF scenarios, and exports a formatted Excel workbook
to the ``output/`` directory.

Dependencies: ``pandas`` and ``openpyxl``.
"""

from __future__ import annotations

import os
from datetime import date
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .data_fetcher import get_fetcher
from .dcf_model import DCFModel

# ---------------------------------------------------------------------------
# Scenario configuration
# ---------------------------------------------------------------------------

_TERMINAL_GROWTH_RATE: float = 0.025
_FORECAST_YEARS: int = 5

# Deltas applied on top of the base case
_BULL_GROWTH_DELTA: float = 0.02
_BULL_WACC_DELTA: float = -0.01
_BEAR_GROWTH_DELTA: float = -0.02
_BEAR_WACC_DELTA: float = 0.01

_SCENARIO_ORDER: tuple[str, ...] = ("Bear Case", "Base Case", "Bull Case")

# ---------------------------------------------------------------------------
# Excel styling constants
# ---------------------------------------------------------------------------

_FMT_CURRENCY = "$#,##0.00"
_FMT_CURRENCY_LARGE = "$#,##0"
_FMT_PERCENT = "0.00%"
_FMT_NUMBER = "#,##0"
_FMT_DECIMAL4 = "0.0000"

_FILL_DARK = PatternFill("solid", fgColor="1F4E79")    # dark navy – title rows
_FILL_SECTION = PatternFill("solid", fgColor="2E75B6")  # medium blue – section headers
_FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")   # pale blue – column headers
_FILL_BEAR = PatternFill("solid", fgColor="FFCCCC")     # light red – bear scenario
_FILL_BASE = PatternFill("solid", fgColor="FFF2CC")     # light amber – base scenario
_FILL_BULL = PatternFill("solid", fgColor="E2EFDA")     # light green – bull scenario
_FILL_TOTAL = PatternFill("solid", fgColor="F2F2F2")    # light grey – totals/subtotals

_SCENARIO_FILLS: dict[str, PatternFill] = {
    "Bear Case": _FILL_BEAR,
    "Base Case": _FILL_BASE,
    "Bull Case": _FILL_BULL,
}

_FONT_TITLE = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
_FONT_SECTION = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_FONT_COL_HDR = Font(name="Calibri", bold=True, size=11)
_FONT_BOLD = Font(name="Calibri", bold=True, size=11)
_FONT_NORMAL = Font(name="Calibri", size=11)
_FONT_ITALIC = Font(name="Calibri", italic=True, size=10, color="595959")

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def run_scenario_analysis(ticker: str, api_key: str | None = None) -> str:
    """Run Bear / Base / Bull DCF scenario analysis and export to Excel.

    Fetches live financial data for *ticker* from the FMP API, derives three
    valuation scenarios from the historical FCF growth rate and current WACC,
    then writes a formatted Excel workbook to
    ``output/{ticker}_DCF_Valuation.xlsx`` (relative to the project root).

    The ``output/`` directory is created automatically if it does not exist.

    Scenarios
    ---------
    * **Base Case** – historical average FCF growth rate, current WACC.
    * **Bull Case** – Base growth + 2 %, WACC − 1 %.
    * **Bear Case** – Base growth − 2 %, WACC + 1 %.

    Parameters
    ----------
    ticker:
        The equity ticker symbol (e.g. ``"AAPL"``).
    api_key:
        A valid FMP API key.  Falls back to the ``FMP_API_KEY`` environment
        variable when ``None``.

    Returns
    -------
    str
        Absolute path to the generated Excel file.

    Raises
    ------
    DataFetchError
        If any upstream API call fails.
    DCFModelError
        If the fetched data produces an invalid DCF configuration (e.g. WACC
        not greater than terminal growth rate).
    """
    ticker = ticker.upper().strip()
    fetcher = get_fetcher(ticker, api_key=api_key)

    # --- Fetch all required financial data ---
    cash_flow_data = fetcher.get_cash_flow_statement()
    enterprise_metrics = fetcher.get_enterprise_metrics()
    wacc_data = fetcher.get_wacc()
    price_data = fetcher.get_current_price()
    base_growth = fetcher.get_historical_fcf_growth()

    wacc = wacc_data["wacc"]

    # --- Define scenario parameters ---
    scenarios: dict[str, dict[str, float]] = {
        "Bear Case": {
            "fcf_growth_rate": base_growth + _BEAR_GROWTH_DELTA,
            "wacc": wacc + _BEAR_WACC_DELTA,
        },
        "Base Case": {
            "fcf_growth_rate": base_growth,
            "wacc": wacc,
        },
        "Bull Case": {
            "fcf_growth_rate": base_growth + _BULL_GROWTH_DELTA,
            "wacc": wacc + _BULL_WACC_DELTA,
        },
    }

    # --- Run DCF model for each scenario ---
    model = DCFModel(cash_flow_data, enterprise_metrics)
    results: dict[str, dict[str, Any]] = {}
    for name, params in scenarios.items():
        r = model.calculate_intrinsic_value(
            fcf_growth_rate=params["fcf_growth_rate"],
            wacc=params["wacc"],
            terminal_growth_rate=_TERMINAL_GROWTH_RATE,
            years=_FORECAST_YEARS,
        )
        # Stash the scenario-level parameters inside the result for convenience
        r["wacc_used"] = params["wacc"]
        r["fcf_growth_rate"] = params["fcf_growth_rate"]
        results[name] = r

    # --- Resolve output directory (project root / output/) ---
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_dir = os.path.join(project_root, "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, f"{ticker}_DCF_Valuation.xlsx")

    # --- Write Excel workbook ---
    _build_excel(output_path, ticker, price_data["price"], scenarios, results)

    return output_path


# ---------------------------------------------------------------------------
# Excel builder
# ---------------------------------------------------------------------------


def _build_excel(
    path: str,
    ticker: str,
    current_price: float,
    scenarios: dict[str, dict[str, float]],
    results: dict[str, dict[str, Any]],
) -> None:
    """Write a formatted multi-sheet Excel workbook to *path*.

    Uses pandas :class:`~pandas.ExcelWriter` (openpyxl engine) to create the
    file and write the FCF projection DataFrames, then re-opens the workbook
    with openpyxl to apply visual formatting, add section headers, and
    populate non-tabular content such as the summary and valuation bridge.
    """
    analysis_date = date.today().isoformat()

    # ---- Phase 1: create workbook skeleton with pandas ----
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Summary sheet – created via an empty frame; populated by openpyxl later
        pd.DataFrame().to_excel(writer, sheet_name="Summary", index=False)

        # Per-scenario sheets – FCF projection table at row 5 (startrow=4)
        for name in _SCENARIO_ORDER:
            df = _make_fcf_df(results[name])
            df.to_excel(writer, sheet_name=name, startrow=4, index=False)

    # ---- Phase 2: enrich with openpyxl ----
    wb = load_workbook(path)

    _populate_summary(
        wb["Summary"], ticker, analysis_date, current_price, scenarios, results
    )
    for name in _SCENARIO_ORDER:
        _populate_scenario(
            wb[name], name, ticker, analysis_date, scenarios[name], results[name]
        )

    wb.save(path)


# ---------------------------------------------------------------------------
# pandas helpers
# ---------------------------------------------------------------------------


def _make_fcf_df(result: dict[str, Any]) -> pd.DataFrame:
    """Build the FCF projection :class:`~pandas.DataFrame` for one scenario."""
    wacc = result["wacc_used"]
    rows = [
        {
            "Year": yr["year"],
            "Projected FCF ($)": yr["projected_fcf"],
            "Discount Factor": 1.0 / (1.0 + wacc) ** yr["year"],
            "Present Value ($)": yr["present_value"],
        }
        for yr in result["projected_fcfs"]
    ]
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# openpyxl sheet population
# ---------------------------------------------------------------------------


def _populate_summary(
    ws: Any,
    ticker: str,
    analysis_date: str,
    current_price: float,
    scenarios: dict[str, dict[str, float]],
    results: dict[str, dict[str, Any]],
) -> None:
    """Write and format the Summary worksheet entirely with openpyxl."""
    # Column layout: A=metric label, B=Bear, C=Base, D=Bull
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.row_dimensions[1].height = 28

    # --- Row 1: workbook title (merged A1:D1) ---
    ws.merge_cells("A1:D1")
    _wc(
        ws, 1, 1,
        f"DCF Valuation Analysis  –  {ticker}",
        font=_FONT_TITLE, fill=_FILL_DARK,
        align=_ALIGN_CENTER,
    )

    # --- Row 2: analysis date ---
    _wc(ws, 2, 1, "Analysis Date:", font=_FONT_BOLD, align=_ALIGN_LEFT)
    _wc(ws, 2, 2, analysis_date, font=_FONT_NORMAL, align=_ALIGN_LEFT)

    # --- Row 4: "Scenario Assumptions" section header (merged) ---
    ws.merge_cells("A4:D4")
    _wc(
        ws, 4, 1, "SCENARIO ASSUMPTIONS",
        font=_FONT_SECTION, fill=_FILL_SECTION, align=_ALIGN_CENTER,
    )
    ws.row_dimensions[4].height = 20

    # --- Row 5: column headers ---
    for col, label in enumerate(["Metric", "Bear Case", "Base Case", "Bull Case"], 1):
        fill = _SCENARIO_FILLS.get(label, _FILL_HEADER)
        _wc(ws, 5, col, label, font=_FONT_COL_HDR, fill=fill, align=_ALIGN_CENTER)

    # --- Rows 6-8: assumptions data ---
    assumption_rows: list[tuple[str, str, Any, Any, Any]] = [
        ("FCF Growth Rate", _FMT_PERCENT,
         scenarios["Bear Case"]["fcf_growth_rate"],
         scenarios["Base Case"]["fcf_growth_rate"],
         scenarios["Bull Case"]["fcf_growth_rate"]),
        ("WACC", _FMT_PERCENT,
         scenarios["Bear Case"]["wacc"],
         scenarios["Base Case"]["wacc"],
         scenarios["Bull Case"]["wacc"]),
        ("Terminal Growth Rate", _FMT_PERCENT,
         _TERMINAL_GROWTH_RATE, _TERMINAL_GROWTH_RATE, _TERMINAL_GROWTH_RATE),
    ]
    for row_idx, (label, fmt, bear_val, base_val, bull_val) in enumerate(
        assumption_rows, start=6
    ):
        _wc(ws, row_idx, 1, label, font=_FONT_NORMAL, align=_ALIGN_LEFT)
        for col, (val, name) in enumerate(
            [(bear_val, "Bear Case"), (base_val, "Base Case"), (bull_val, "Bull Case")],
            start=2,
        ):
            _wc(
                ws, row_idx, col, val,
                font=_FONT_NORMAL, fill=_SCENARIO_FILLS[name],
                align=_ALIGN_RIGHT, fmt=fmt,
            )

    # --- Row 10: "Valuation Summary" section header (merged) ---
    ws.merge_cells("A10:D10")
    _wc(
        ws, 10, 1, "VALUATION SUMMARY",
        font=_FONT_SECTION, fill=_FILL_SECTION, align=_ALIGN_CENTER,
    )
    ws.row_dimensions[10].height = 20

    # --- Row 11: column headers ---
    for col, label in enumerate(["Metric", "Bear Case", "Base Case", "Bull Case"], 1):
        fill = _SCENARIO_FILLS.get(label, _FILL_HEADER)
        _wc(ws, 11, col, label, font=_FONT_COL_HDR, fill=fill, align=_ALIGN_CENTER)

    # --- Rows 12-20: valuation data ---
    valuation_rows: list[tuple[str, str, Any, Any, Any]] = [
        ("Enterprise Value", _FMT_CURRENCY_LARGE,
         results["Bear Case"]["enterprise_value"],
         results["Base Case"]["enterprise_value"],
         results["Bull Case"]["enterprise_value"]),
        ("(-) Total Debt", _FMT_CURRENCY_LARGE,
         results["Bear Case"]["total_debt"],
         results["Base Case"]["total_debt"],
         results["Bull Case"]["total_debt"]),
        ("(+) Cash", _FMT_CURRENCY_LARGE,
         results["Bear Case"]["cash"],
         results["Base Case"]["cash"],
         results["Bull Case"]["cash"]),
        ("Equity Value", _FMT_CURRENCY_LARGE,
         results["Bear Case"]["equity_value"],
         results["Base Case"]["equity_value"],
         results["Bull Case"]["equity_value"]),
        ("Shares Outstanding", _FMT_NUMBER,
         results["Bear Case"]["shares_outstanding"],
         results["Base Case"]["shares_outstanding"],
         results["Bull Case"]["shares_outstanding"]),
        ("Intrinsic Value / Share", _FMT_CURRENCY,
         results["Bear Case"]["intrinsic_value_per_share"],
         results["Base Case"]["intrinsic_value_per_share"],
         results["Bull Case"]["intrinsic_value_per_share"]),
        ("Current Stock Price", _FMT_CURRENCY,
         current_price, current_price, current_price),
        ("Upside / (Downside)", _FMT_PERCENT,
         _upside(results["Bear Case"]["intrinsic_value_per_share"], current_price),
         _upside(results["Base Case"]["intrinsic_value_per_share"], current_price),
         _upside(results["Bull Case"]["intrinsic_value_per_share"], current_price)),
    ]
    _BOLD_ROWS = {"Enterprise Value", "Equity Value", "Intrinsic Value / Share"}
    for row_idx, (label, fmt, bear_val, base_val, bull_val) in enumerate(
        valuation_rows, start=12
    ):
        font = _FONT_BOLD if label in _BOLD_ROWS else _FONT_NORMAL
        _wc(ws, row_idx, 1, label, font=font, align=_ALIGN_LEFT)
        for col, (val, name) in enumerate(
            [(bear_val, "Bear Case"), (base_val, "Base Case"), (bull_val, "Bull Case")],
            start=2,
        ):
            cell_font = _FONT_BOLD if label in _BOLD_ROWS else _FONT_NORMAL
            _wc(
                ws, row_idx, col, val,
                font=cell_font, fill=_SCENARIO_FILLS[name],
                align=_ALIGN_RIGHT, fmt=fmt,
            )


def _populate_scenario(
    ws: Any,
    name: str,
    ticker: str,
    analysis_date: str,
    params: dict[str, float],
    result: dict[str, Any],
) -> None:
    """Write title, format the pandas FCF table, and add the valuation bridge."""
    scenario_fill = _SCENARIO_FILLS[name]

    # --- Column widths ---
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 25
    ws.row_dimensions[1].height = 24

    # --- Row 1: sheet title (merged A1:D1) ---
    ws.merge_cells("A1:D1")
    _wc(
        ws, 1, 1,
        f"{name}  –  {ticker}",
        font=_FONT_TITLE, fill=_FILL_DARK, align=_ALIGN_CENTER,
    )

    # --- Row 2: scenario parameters ---
    ws.merge_cells("A2:D2")
    param_text = (
        f"FCF Growth: {params['fcf_growth_rate']:.2%}  |  "
        f"WACC: {params['wacc']:.2%}  |  "
        f"Terminal Growth: {_TERMINAL_GROWTH_RATE:.2%}  |  "
        f"Analysis Date: {analysis_date}"
    )
    _wc(ws, 2, 1, param_text, font=_FONT_ITALIC, align=_ALIGN_CENTER)

    # --- Row 4: "FCF Projections" section header (merged A4:D4) ---
    ws.merge_cells("A4:D4")
    _wc(
        ws, 4, 1, "FCF PROJECTIONS",
        font=_FONT_SECTION, fill=_FILL_SECTION, align=_ALIGN_CENTER,
    )
    ws.row_dimensions[4].height = 20

    # --- Rows 5-10: pandas wrote the FCF table here; apply formatting ---
    # Row 5 = column headers (Year | Projected FCF ($) | Discount Factor | Present Value ($))
    for col in range(1, 5):
        cell = ws.cell(row=5, column=col)
        cell.font = _FONT_COL_HDR
        cell.fill = scenario_fill
        cell.alignment = _ALIGN_CENTER

    # Rows 6-10 = data
    col_formats = {1: _FMT_NUMBER, 2: _FMT_CURRENCY_LARGE, 3: _FMT_DECIMAL4, 4: _FMT_CURRENCY_LARGE}
    for data_row in range(6, 6 + _FORECAST_YEARS):
        for col, fmt in col_formats.items():
            cell = ws.cell(row=data_row, column=col)
            cell.font = _FONT_NORMAL
            cell.number_format = fmt
            cell.alignment = _ALIGN_RIGHT

    # --- Row 12: "Valuation Bridge" section header ---
    ws.merge_cells("A12:D12")
    _wc(
        ws, 12, 1, "VALUATION BRIDGE",
        font=_FONT_SECTION, fill=_FILL_SECTION, align=_ALIGN_CENTER,
    )
    ws.row_dimensions[12].height = 20

    # --- Rows 13-21: valuation bridge (label + value, columns A-B) ---
    bridge_rows: list[tuple[str, Any, str, bool]] = [
        ("Sum of PV(FCFs)", result["sum_pv_fcfs"], _FMT_CURRENCY_LARGE, False),
        ("Terminal Value", result["terminal_value"], _FMT_CURRENCY_LARGE, False),
        ("PV of Terminal Value", result["pv_terminal_value"], _FMT_CURRENCY_LARGE, False),
        ("(=) Enterprise Value", result["enterprise_value"], _FMT_CURRENCY_LARGE, True),
        ("(-) Total Debt", result["total_debt"], _FMT_CURRENCY_LARGE, False),
        ("(+) Cash", result["cash"], _FMT_CURRENCY_LARGE, False),
        ("(=) Equity Value", result["equity_value"], _FMT_CURRENCY_LARGE, True),
        ("Shares Outstanding", result["shares_outstanding"], _FMT_NUMBER, False),
        ("(=) Intrinsic Value / Share", result["intrinsic_value_per_share"], _FMT_CURRENCY, True),
    ]
    for row_idx, (label, value, fmt, is_total) in enumerate(bridge_rows, start=13):
        font = _FONT_BOLD if is_total else _FONT_NORMAL
        fill = _FILL_TOTAL if is_total else None
        _wc(ws, row_idx, 1, label, font=font, fill=fill, align=_ALIGN_LEFT)
        _wc(ws, row_idx, 2, value, font=font, fill=fill, align=_ALIGN_RIGHT, fmt=fmt)
        # Merge unused columns to keep it clean
        ws.merge_cells(
            start_row=row_idx, start_column=3,
            end_row=row_idx, end_column=4,
        )


# ---------------------------------------------------------------------------
# Low-level openpyxl helper
# ---------------------------------------------------------------------------


def _upside(ivps: float, current_price: float) -> float:
    """Return the percentage upside of *ivps* relative to *current_price*."""
    return (ivps / current_price) - 1.0 if current_price else 0.0


def _wc(
    ws: Any,
    row: int,
    col: int,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    align: Alignment | None = None,
    fmt: str | None = None,
) -> Any:
    """Write a cell value and optionally set font, fill, alignment, and format."""
    cell = ws.cell(row=row, column=col, value=value)
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if align is not None:
        cell.alignment = align
    if fmt is not None:
        cell.number_format = fmt
    return cell
